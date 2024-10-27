import sys
import Database
import TableModel
import openpyxl
import os
from reportlab.pdfgen import canvas

from pathlib import Path

from PyQt5 import QtWidgets

from PyQt5.QtWidgets import (
    QAbstractItemView,
    QApplication,
    QDialog,
    QMainWindow,
    QMessageBox,
)

from PyQt5.QtGui import (
    QColor,
    QPixmap,
    QIcon,
    QRegExpValidator,
    QValidator,
)

from PyQt5.QtWidgets import QApplication  

from PyQt5 import QtCore

from PyQt5.uic import loadUi



class PartyWindow(QMainWindow):  
    """Main Window for the Party List application."""  
  
    def __init__(self, parent=None):  
        """
        Initializer for the PartyWindow class.

        - Initializes the parent class (QMainWindow).
        - Creates an instance of the PersonModel class to manage the table data.
        - Loads the UI from a `.ui` file using PyQt's loadUi function.

        - Sets up the table view:
            - Sets the model for the table view to the model instance.
            - Sets the selection behavior to select entire rows on click.
            - Applies custom styles to the horizontal and vertical headers
                using Qt stylesheets.

        - Connects button clicks to corresponding methods for:
            - Adding a new person.
            - Deleting the selected person.
            - Deleting all people.
            - Closing the application.
            - Exporting the selected row to Excel.
            - Exporting the entire party list to Excel.
        """

        super().__init__(parent)
        self.personModel = TableModel.PersonModel()
        loadUi("PartyList.ui", self)

        dirname = os.path.dirname(__file__)
        self.warning = os.path.join(dirname, 'warning.png')
        self.info = os.path.join(dirname, 'info.png')

        self.table_party_list.setModel(self.personModel.model)
        self.table_party_list.setSelectionBehavior(QAbstractItemView.SelectRows) 
        self.table_party_list.horizontalHeader().setStyleSheet('QHeaderView::section {background-color: #4562C3; color: rgb(250, 250, 250); font-weight: bold;} QToolTip { background-color: #8ad4ff; color: black; border: #8ad4ff solid 1px}')
        self.table_party_list.verticalHeader().setStyleSheet('QHeaderView::section {background-color: rgb(60, 60, 60); color: rgb(200, 250, 200); font-size: 6} QToolTip { background-color: #8ad4ff; color: black; border: #8ad4ff solid 1px}')

        self.button_new.clicked.connect(self.openAddPerson)  
        self.button_delete.clicked.connect(lambda: self.deletePerson())
        self.button_delete_all.clicked.connect(lambda: self.deletePeople())
        self.button_close.clicked.connect(app.closeAllWindows)
        self.button_export_one.clicked.connect(self.export_selected_row)
        self.button_export_all_excel.clicked.connect(self.export_to_excel)
        self.button_export_all_pdf.clicked.connect(self.export_to_pdf) 


    def deletePerson(self):  
        """
        Deletes the selected person from the database.

        - Gets the index of the currently selected row.
        - Checks if a row is selected (index >= 0).
        - Displays a confirmation message box with a warning about deleting.
        - If the user confirms deletion (QMessageBox.Ok):
            - Calls the deletePerson method of the PersonModel instance to remove
              the person from the database.
        """

        row = self.table_party_list.currentIndex().row()  
        if row < 0:  
            return  

        messagebox = QMessageBox(QMessageBox.Information, "Warning!", "<FONT COLOR='white'>Really do you want to delete the selected person?", QMessageBox.Ok | QMessageBox.Cancel, parent=self)
        messagebox.setIconPixmap(QPixmap(self.warning))
        result = messagebox.exec_()  # Store the return value

        if result == QMessageBox.Ok:
            self.personModel.deletePerson(row)

    def deletePeople(self):  
        """
        Deletes all people from the database.

        - Displays a confirmation message box with a warning about deleting all people.
        - If the user confirms deletion (QMessageBox.Ok):
            - Calls the `deletePeople` method of the `PersonModel` instance to remove
            all people from the database.
        """
        messagebox = QMessageBox(QMessageBox.Information, "Warning!", "<FONT COLOR='white'>Really do you want to delete all people?", QMessageBox.Ok | QMessageBox.Cancel, parent=self)
        messagebox.setIconPixmap(QPixmap(self.warning))
        result = messagebox.exec_()  # Store the return value

        if result == QMessageBox.Ok:
            self.personModel.deletePeople() 

    def openAddPerson(self):  
        """
        Open the Add Contact dialog.

        - Creates an instance of the `AddPersonDialog` class.
        - Displays the dialog using `dialog.exec()`.
        - If the dialog is accepted (user clicks "Ok"):
            - Retrieves the data entered in the dialog using the `dialog.data` attribute (assumed to be a list).
            - Calls the `addPerson` method of the `PersonModel` instance to add the new person's data to the database.
            - Resizes the table columns to fit the content automatically using `self.table_party_list.resizeColumnsToContents()`.
        """

        dialog = AddPersonDialog()  
        if dialog.exec() == QDialog.Accepted:  
            self.personModel.addPerson(dialog.data)  

    def export_to_excel(self):
        """
        Exports party list to an Excel file using retrieved dictionaries.

        - Retrieves all people data using `self.personModel.get_all_people()`.
        - Checks if there's any data to export:
            - If no data is found, displays an information message box and returns.
        - Creates a new Excel workbook using `openpyxl.Workbook()`.
        - Sets the active sheet title to "Party List" using `sheet.title = "Party List"`.

        - Writes headers:
            - Assumes a consistent data structure in dictionaries.
            - Extracts headers from the first dictionary's keys (`list(people_data[0].keys())`).
            - Iterates through headers and writes them to the first row (row=1) of the sheet.

        - Writes data:
            - Iterates through each person's data dictionary (`person`):
                - Iterates through each key-value pair (`value`) in the dictionary.
                - Writes the value to the corresponding cell in the sheet based on row and column indices.
                - Starts writing data from the second row (row=2) to avoid overwriting headers.

        - Saves the workbook:
            - Prompts the user to choose a filename using `QtWidgets.QFileDialog.getSaveFileName`.
            - If a filename is chosen, saves the workbook using `wb.save(filename)`.
            - Displays a success message box using `QMessageBox.information` with white font color for better visibility.
        """

        people_data = self.personModel.get_all_people()

        if not people_data:
            messagebox = QMessageBox(QMessageBox.Information, "Information", "<FONT COLOR='white'>No data found to export!", QMessageBox.Ok, parent=self)
            messagebox.setIconPixmap(QPixmap(self.info))
            result = messagebox.exec_()  # Store the return value
            return

        # Create a new workbook
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = "Party List"

        # Write headers
        headers = list(people_data[0].keys())  # Assuming consistent data structure
        for col_index, header in enumerate(headers, start=1):
            sheet.cell(row=1, column=col_index).value = header

        # Write data
        for row_index, person in enumerate(people_data, start=2):
            for col_index, value in enumerate(person.values(), start=1):
                sheet.cell(row=row_index, column=col_index).value = value

        # Save the workbook
        filename, _ = QtWidgets.QFileDialog.getSaveFileName(self, "Export Party List", "", "Excel Files (*.xlsx)")
        if filename:
            wb.save(filename)
            messagebox = QMessageBox(QMessageBox.Information, "Information", "<FONT COLOR='white'>Party list exported to Excel successfully!", QMessageBox.Ok, parent=self)
            messagebox.setIconPixmap(QPixmap(self.info))
            result = messagebox.exec_()  # Store the return value

    def export_selected_row(self):
        """
        Exports the data from the currently selected row to an Excel file.

        - Gets the index of the currently selected row using `self.table_party_list.currentIndex().row()`.
        - Checks if a row is selected:
            - If no row is selected (index < 0), returns without further action.
        - Retrieves data from the selected row using `self.personModel.get_current_row(row)`.
        - Checks if any data is retrieved:
            - If no data is found for the selected row, displays an information message box and returns.

        - Creates a new Excel workbook (same as `export_to_excel`).
        - Sets the active sheet title (same as `export_to_excel`).

        - Writes headers:
            - Similar to `export_to_excel`, assumes a consistent data structure and extracts headers from the retrieved data's keys.
            - Writes headers to the first row of the sheet.

        - Writes data:
            - Iterates through the key-value pairs in the retrieved data dictionary.
            - Writes values to corresponding cells in the sheet, starting from the second row.

        - Saves the workbook (same as `export_to_excel`).
        - Displays a success message box (same as `export_to_excel`).
        """
        row = self.table_party_list.currentIndex().row()  
        if row < 0:  
            return 

        selected_data = self.personModel.get_current_row(row)  # Get data from selected row

        if not selected_data:
            messagebox = QMessageBox(QMessageBox.Information, "Information", "<FONT COLOR='white'>No row selected to export!", QMessageBox.Ok, parent=self)
            messagebox.setIconPixmap(QPixmap(self.info))
            result = messagebox.exec_()  # Store the return value
            return

        # Create a new workbook
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = "Party List"

        # Write headers (assuming consistent data structure)
        headers = list(selected_data.keys())
        for col_index, header in enumerate(headers, start=1):
            sheet.cell(row=1, column=col_index).value = header

        # Write data
        row_index = 2  # Start data on row 2
        for col_index, value in enumerate(selected_data.values(), start=1):
            sheet.cell(row=row_index, column=col_index).value = value

        # Save the workbook
        filename, _ = QtWidgets.QFileDialog.getSaveFileName(self, "Export Selected Row", "", "Excel Files (*.xlsx)")
        if filename:
            wb.save(filename)
            messageBox = QMessageBox.information(self, "Success", "<FONT COLOR='white'>Selected row exported to Excel successfully!</FONT>")

    def export_to_pdf(self):
        """
        Exports party list to a PDF file using retrieved dictionaries.

        - Retrieves all people data using `self.personModel.get_all_people()`.
        - Checks if there's any data to export:
            - If no data is found, displays an information message box and returns.
        - Creates a new PDF document using `canvas.Canvas`.
        - Sets document title and font properties.

        - Writes headers:
            - Assumes a consistent data structure in dictionaries.
            - Extracts headers from the first dictionary's keys.
            - Sets font size and color for headers.
            - Writes headers centered in columns.

        - Writes data:
            - Iterates through each person's data dictionary.
            - Sets font size and color for data.
            - Writes data left-aligned in columns with some margin.

        - Saves the PDF:
            - Prompts the user to choose a filename using `QtWidgets.QFileDialog.getSaveFileName`.
            - If a filename is chosen, saves the document using `canvas.save`.
            - Displays a success message box.
        """

        people_data = self.personModel.get_all_people()

        if not people_data:
            messageBox = QMessageBox.information(self, "Information", "<FONT COLOR='white'>No data found to export!</FONT>")
            return

        # Create a new PDF document
        filename, _ = QtWidgets.QFileDialog.getSaveFileName(self, "Export Party List", "", "PDF Files (*.pdf)")
        if not filename:
            return

        my_canvas = canvas.Canvas(filename)  # Create canvas inside the function
        my_canvas.setTitle("Party List")
        my_canvas.setFont("Helvetica", 12)  # Set font properties

        # Header row properties
        header_y = 780  # Starting position for headers
        header_font_size = 11
        header_color = "blue"
        header_margin = 50

        # Data row properties
        data_y = header_y - 20  # Starting position for data (below headers)
        data_font_size = 11
        data_color = "black"
        data_margin = 50  # Margin from left edge for data

        # Write headers
        headers = list(people_data[0].keys())
        num_columns = len(headers)
        column_width = (500 - data_margin) / num_columns

        for col_index, header in enumerate(headers):
            x = header_margin + (column_width * col_index)
            my_canvas.setFontSize(header_font_size)
            my_canvas.setFillColor(header_color)
            my_canvas.drawCentredString(x, header_y, header)

        # Write data
        for row_index, person in enumerate(people_data):
            data_y -= 15  # Move down for each data row
            my_canvas.setFontSize(data_font_size)
            my_canvas.setFillColor(data_color)

            for col_index, value in enumerate(person.values()):
                x = data_margin + (column_width * col_index)
                my_canvas.drawString(x, data_y, str(value))  # Left-aligned data

        try:
            my_canvas.save()
            messageBox = QMessageBox.information(self, "Success", "<FONT COLOR='white'>Party list exported to PDF successfully!</FONT>")
        except PermissionError as e:
            messageBox = QMessageBox.critical(self, "Error", f"<FONT COLOR='white'>Error saving PDF!</FONT>")

        

# second dialog for informatin display
class AddPersonDialog(QDialog):

    def __init__(self):
        """
        Initializer for the AddPersonDialog class.

        - Initializes the parent class (QDialog).
        - Loads the UI from a `.ui` file using PyQt's loadUi function.
        - Sets the window icon using a QIcon object with the path to an image file.

        - Connects button box signals (accepted/rejected) to respective methods:
            - `accept`: Handles data validation and closing the dialog on OK.
            - `reject`: Closes the dialog on Cancel.
        """

        super().__init__()
        # loading design - created in QtCreator
        loadUi("AddPerson.ui", self)

        self.setWindowIcon(QIcon('Party.png'))

        # button box created in Qt Designer
        # press button OK
        self.buttonBox.accepted.connect(self.accept)  
        # press butoon Cancel
        self.buttonBox.rejected.connect(self.reject)

    # press button OK
    def accept(self):  
        """
        Accept the data provided through the dialog and validate required fields.

        - Initializes an empty list `self.data` to store collected data.
        - Iterates through input fields (first name, last name, phone, email).
        - For each field:
            - Checks if the field is empty using `field.text()`.
            - If empty:
                - Displays a critical message box with the missing field name
                  using QMessageBox.critical.
                - Resets `self.data` to None to prevent partial data.
                - Returns to avoid further processing.
        - If all fields have data, appends the text from each field to `self.data`.
        - Calls the superclass `accept` method to close the dialog with OK.
        """
        dirname = os.path.dirname(__file__)
        stop_writing = os.path.join(dirname, 'stop_writing.png')

        self.data = []  
        for field in (self.edit_first_name, self.edit_last_name, self.edit_phone, self.edit_email):  
            if not field.text():  
                if field == self.edit_first_name:
                    label_text = self.label_first_name.text()
                elif field == self.edit_last_name:
                    label_text = self.label_last_name.text()
                elif field == self.edit_phone:
                    label_text = self.label_phone.text()
                elif field == self.edit_email:
                    label_text = self.label_email.text()

                messagebox = QMessageBox(QMessageBox.Information, "Error", f"<FONT COLOR='white'>You must provide a person's {label_text}", buttons=QMessageBox.Ok, parent=self)
                messagebox.setIconPixmap(QPixmap(stop_writing))
                messagebox.exec_()


                self.data = None  # Reset .data  
                return  

            self.data.append(field.text())  
            print(self.data)

        super().accept()  


app = QApplication(sys.argv)
connection = Database.createConnection("party_list.db")
mainwindow = PartyWindow()
widget = QtWidgets.QStackedWidget()
widget.setWindowTitle('PARTY LIST')
widget.setWindowIcon(QIcon('Party.png'))
widget.addWidget(mainwindow)
widget.show()
app.setStyleSheet(Path('styles.qss').read_text())
app.exec()